from openpyxl import load_workbook
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.metrics import precision_recall_fscore_support
import numpy as np
import sklearn

train_jobs = []
train_tags = []
tfidf = ""

# Loading data and formatting it from excel sheet.
def loadData():

    global train_jobs
    global train_tags

    wb = load_workbook("FinalGlassdoorJobs-869.xlsx")
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(2, sheet.max_row):
        list = []
        a = sheet.cell(row=i, column=3).value
        if a != None:
            train_jobs.append(a);

            if sheet.cell(row=i, column=4).value == 1:
                list.append(0);
            if sheet.cell(row=i, column=5).value == 1:
                list.append(1);
            if sheet.cell(row=i, column=8).value == 1:
                list.append(2);
            if sheet.cell(row=i, column=9).value == 1:
                list.append(3);

            train_tags.append(list)

# It is the text procesor tf_idf
# Finds signature words and computes individuals score.
def tf_idf():
    global tfidf
    global train_jobs

    tfidf = TfidfVectorizer(tokenizer=tokenize, min_df=3, max_df=0.90, max_features=2000, use_idf=True, sublinear_tf=True)
    tfidf.fit(train_jobs)

# Tokenizer used by tfidf
def tokenize(text):
    #cachedStopWords = stopwords.words("english")
    tokens = nltk.word_tokenize(text)
    stems = []

    for item in tokens:
        stems.append(PorterStemmer().stem(item))
    return stems

# collects training set
def get_train_set():
    global tfidf
    global train_jobs
    global train_tags

    return tfidf.transform(train_jobs), MultiLabelBinarizer().fit_transform(train_tags)

# Finds k nearest jobs of provided job using euclidean distance.
def KNN(jobs, predict, k=3):
    distances = []
    for i in range(0,len(jobs)):
        distance = np.linalg.norm(np.array(jobs[i])-np.array(predict))
        distances.append([distance, i])
    votes = [i for i in sorted(distances, key = getKey)[:k]]
    return votes

def getKey(item):
    return item[0]

# It computes the confidence value for every tags for the provided job.
def MAP(similarJobs, y_train, k):    
    predictedTagsCounter = [0,0,0,0]
    
    for i in similarJobs:
        jobTag = y_train[i[1]]
        for j in range(0,4):
            if(jobTag[j] == 1):
                predictedTagsCounter[j] += 1
                
    for i in range(0,4):
        if(predictedTagsCounter[i]/k >= 0.70):
        # if (predictedTagsCounter[i]/k > confidence[i]):
            predictedTagsCounter[i] = 1
        else:
            predictedTagsCounter[i] = 0
    
    return (predictedTagsCounter)

# called by api to predicts tags
def predictByKNN(job):

    loadData()
    tf_idf()

    X, Y = get_train_set()
    y_test_predicted = []
    x_train, y_train = X.toarray()[:200], Y[:200]
    # x_test, y_test = X.toarray()[175:], Y[175:]

    return prediction(x_train, y_train, job)

# Uses the models and predicts the a set of tags using confidence value.
def prediction(x_train, y_train, job):
    
    global tfidf
    
    features = tfidf.transform([job])
    features = features.toarray()[0]
    
    k = 5
    similarJob = KNN(x_train, features, k)
    predicted = MAP(similarJob, y_train, k)
    return predicted