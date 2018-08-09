from openpyxl import load_workbook
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.metrics import precision_recall_fscore_support
import numpy as np
import sklearn
import tensorflow as tf

from keras.models import Sequential
from keras.layers import Activation, Dense

train_jobs = []
train_tags = []
tfidf = ""


def loadData():

    global train_jobs
    global train_tags

    wb = load_workbook("FinalGlassdoorJobs-869.xlsx")
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(2, sheet.max_row):
        list = []
        a = sheet.cell(row=i, column=3).value
        if a != None:
            train_jobs.append(a);

            if sheet.cell(row=i, column=4).value == 1:
                list.append(0);
            if sheet.cell(row=i, column=5).value == 1:
                list.append(1);
            if sheet.cell(row=i, column=8).value == 1:
                list.append(2);
            if sheet.cell(row=i, column=9).value == 1:
                list.append(3);

            train_tags.append(list)

def tf_idf():
    global tfidf
    global train_jobs

    tfidf = TfidfVectorizer(tokenizer=tokenize, min_df=3, max_df=0.90, max_features=2000, use_idf=True, sublinear_tf=True)
    tfidf.fit(train_jobs)

def tokenize(text):
    #cachedStopWords = stopwords.words("english")
    tokens = nltk.word_tokenize(text)
    stems = []

    for item in tokens:
        stems.append(PorterStemmer().stem(item))
    return stems

def get_train_set():
    global tfidf
    global train_jobs
    global train_tags

    return tfidf.transform(train_jobs), MultiLabelBinarizer().fit_transform(train_tags)


def checkAccuracy(y_test, y_test_predicted):
    TP = 0
    TN = 0
    FN = 0
    FP = 0

    for i in range(0,len(y_test)):
        for j in range(0,4):
            if (y_test[i][j] == 1 and y_test_predicted[i][j] == 1):
                TP += 1
            if (y_test[i][j] == 0 and y_test_predicted[i][j] == 0):
                TN += 1
            if (y_test[i][j] == 1 and y_test_predicted[i][j] == 0):
                FN += 1
            if (y_test[i][j] == 0 and y_test_predicted[i][j] == 1):
                FP += 1
    # https://www.kdnuggets.com/faq/precision-recall.html [Computation of accuracy, precision and recall]
    accuracy = (TN+TP)/(TN+TP+FN+FP) 
    precision = (TP)/(TP+FP) 
    recall = (TP)/(TP+FN) 
    f1 = 2 * (recall * precision) / (recall + precision) 
    
    return accuracy, precision, recall, f1

def nnModel(x_train, y_train):
    model = Sequential()
    # Input Layer
    model.add(Dense(100, input_dim=len(x_train[0]), activation='relu'))

    # Hidden Layer/s
    model.add(Dense(200, activation='relu'))
    # model.add(Dense(50, activation='relu'))

    # Output Layer - bernoulli distribution
    model.add(Dense(4, activation='sigmoid'))

    model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])

    history = model.fit(x_train, y_train, epochs=10, batch_size=10)

    # score = model.evaluate(x_test, y_test)

    return model


def trainNeuralNetwork():

    global tfidf

    loadData()
    tf_idf()

    X, Y = get_train_set()
    y_test_predicted = []
    x_train, y_train = X.toarray()[:200], Y[:200]
    # x_test, y_test = X.toarray()[175:], Y[175:]

    return nnModel(x_train, y_train), tfidf, tf.get_default_graph()

def predictByNN(tfidf, model, graph, job):

    features = tfidf.transform([job])

    predictions = [0,0,0,0]
    with graph.as_default():
        predictions = model.predict(features)
    
    rounded = [(int)(round(p)) for p in predictions[0]]
    
    return rounded